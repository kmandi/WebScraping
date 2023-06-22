import openpyxl
import requests
from bs4 import BeautifulSoup

ex = openpyxl.Workbook()
print(ex.sheetnames)
sheet = ex.active
sheet.title ='Top Rated Movies'
print(ex.sheetnames)
sheet.append('Movie Rank', 'Movie Name', 'Year of Release', 'IMDB Rating' )

try:
    source = requests.get('https://www.imdb.com/chart/toptv/')
    source.raise_for_status()

    soup = BeautifulSoup(source.text, 'html.parser')
    movies = soup.find('tbody', class_='lister-list').find_all('tr')

    for movie in movies:
        rank = movie.find('td', class_='titleColumn').get_text(strip=True)[:1]
        name = movie.find('td', class_='titleColumn').a.text
        year = movie.find('td', class_='titleColumn').span.text.strip('()')
        rating = movie.find('td', class_='ratingColumn').strong.text
        print(rank, name, year, rating)
        sheet.append([rank, name, year, rating])

except Exception as e:
    print(e)